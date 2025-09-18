# app.py ‚Äî Pareto Comunidad (DELITO / RIESGO SOCIAL / OTROS FACTORES)
# Lee columnas AI:ET y cuenta celdas con datos (no vac√≠as / no "no" / no "0").
# Si detecta encabezados AI:ET con color amarillo, usa solo esas columnas; si no, usa todas AI:ET.

import re
import unicodedata
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from difflib import get_close_matches

st.set_page_config(page_title="Pareto Comunidad ‚Äì MSP", layout="wide", initial_sidebar_state="collapsed")
TOP_N_GRAFICO = 60

# ===================== CATEGOR√çAS PERMITIDAS =====================
CATEGORIAS_VALIDAS = {"DELITO", "RIESGO SOCIAL", "OTROS FACTORES"}

def _force_cat(x: str) -> str:
    n = (x or "").strip().upper()
    return n if n in CATEGORIAS_VALIDAS else "OTROS FACTORES"

# ===================== Diccionario embebido (AMPLIABLE) =====================
DICCIONARIO_EMBEBIDO = pd.DataFrame([
    # -------- DELITO --------
    ["Venta de drogas", "DELITO"],
    ["Consumo de drogas", "DELITO"],
    ["Bunker (Puntos de venta y consumo de drogas)", "DELITO"],
    ["Hurto", "DELITO"],
    ["Robo a personas", "DELITO"],
    ["Robo a vivienda (Tacha)", "DELITO"],
    ["Robo a vivienda (Intimidaci√≥n)", "DELITO"],
    ["Robo a veh√≠culos (Tacha)", "DELITO"],
    ["Robo a comercio (Intimidaci√≥n)", "DELITO"],
    ["Robo a comercio (Tacha)", "DELITO"],
    ["Robo de veh√≠culos", "DELITO"],
    ["Receptaci√≥n", "DELITO"],
    ["Estafas o defraudaci√≥n", "DELITO"],
    ["Da√±os/Vandalismo", "DELITO"],
    ["Lesiones", "DELITO"],
    ["Contrabando", "DELITO"],
    ["Homicidios", "DELITO"],
    ["Extorsi√≥n", "DELITO"],
    ["Delitos sexuales", "DELITO"],
    ["Estafa inform√°tica", "DELITO"],
    ["Robo de cable", "DELITO"],
    ["Robo de bienes agr√≠cola", "DELITO"],
    ["Robo a edificacion (Tacha)", "DELITO"],
    ["Robo a transporte p√∫blico con Intimidaci√≥n", "DELITO"],
    # ---- RIESGO SOCIAL ----
    ["Falta de inversion social", "RIESGO SOCIAL"],
    ["Falta de oportunidades laborales.", "RIESGO SOCIAL"],
    ["Personas con exceso de tiempo de ocio", "RIESGO SOCIAL"],
    ["Violencia intrafamiliar", "RIESGO SOCIAL"],
    ["Desvinculaci√≥n escolar", "RIESGO SOCIAL"],
    ["Abandono de personas (Menor de edad, adulto mayor o capacidades diferentes)", "RIESGO SOCIAL"],
    # ---- OTROS FACTORES ----
    ["Consumo de alcohol en v√≠a p√∫blica", "OTROS FACTORES"],
    ["Deficiencia en la infraestructura vial", "OTROS FACTORES"],
    ["Contaminacion Sonica", "OTROS FACTORES"],
    ["Lotes bald√≠os.", "OTROS FACTORES"],
    ["Falta de salubridad publica", "OTROS FACTORES"],
    ["Disturbios(Ri√±as)", "OTROS FACTORES"],
    ["Personas en situaci√≥n de calle.", "OTROS FACTORES"],
    ["Usurpacion de terrenos (Precarios)", "OTROS FACTORES"],
    ["P√©rdida de espacios p√∫blicos", "OTROS FACTORES"],
    ["Deficiencias en el alumbrado publico", "OTROS FACTORES"],
    ["Acoso sexual callejero", "OTROS FACTORES"],
    ["Hospedajes ilegales (Cuarter√≠as)", "OTROS FACTORES"],
    ["Ventas informales (Ambulantes)", "OTROS FACTORES"],
    ["Maltrato animal", "OTROS FACTORES"],
    ["Tala ilegal", "OTROS FACTORES"],
    ["Trata de personas", "OTROS FACTORES"],
    ["Explotacion Laboral infantil", "OTROS FACTORES"],
    ["Caza ilegal", "OTROS FACTORES"],
    ["Abigeato (Robo y destace de ganado)", "OTROS FACTORES"],
    ["Zona de prostituci√≥n", "OTROS FACTORES"],
    ["Explotacion Sexual infantil", "OTROS FACTORES"],
    ["Tr√°fico ilegal de personas", "OTROS FACTORES"],
    ["Robo de combustible", "OTROS FACTORES"],
    ["Pesca ilegal", "OTROS FACTORES"],
], columns=["Descriptor", "Categor√≠a"]).assign(Categor√≠a=lambda d: d["Categor√≠a"].map(_force_cat))

# ===================== Sin√≥nimos / variaciones (mapeo encabezados) =====================
SINONIMOS: Dict[str, List[str]] = {
    "Venta de drogas": ["venta de drogas", "puntos de venta", "narcomenudeo"],
    "Consumo de drogas": ["consumo de drogas", "consumen drogas", "consumo marihuana", "fumando piedra"],
    "Bunker (Puntos de venta y consumo de drogas)": ["bunker", "bunquer", "b√∫nker"],
    "Hurto": ["hurto", "sustraccion"],
    "Robo a personas": ["asalto a persona", "atraco a persona"],
    "Robo a vivienda (Tacha)": ["tacha vivienda", "robo vivienda tacha"],
    "Robo a vivienda (Intimidaci√≥n)": ["asalto a vivienda", "intimidacion vivienda"],
    "Robo a veh√≠culos (Tacha)": ["tacha vehiculos", "robo tacha vehiculo"],
    "Robo a comercio (Intimidaci√≥n)": ["asalto a comercio"],
    "Robo a comercio (Tacha)": ["robo comercio tacha"],
    "Da√±os/Vandalismo": ["da√±os", "vandalismo", "grafiti", "da√±o a la propiedad"],
    "Receptaci√≥n": ["receptacion", "compra de robado", "reduccion"],
    "Estafas o defraudaci√≥n": ["estafas", "defraudacion", "estafa"],
    "Robo de veh√≠culos": ["robo de vehiculos"],
    "Lesiones": ["golpiza", "lesionados"],
    "Consumo de alcohol en v√≠a p√∫blica": ["licores en via publica"],
    "Deficiencia en la infraestructura vial": ["huecos", "baches", "infraestructura vial"],
    "P√©rdida de espacios p√∫blicos": ["perdida espacios publicos"],
    "Deficiencias en el alumbrado publico": ["alumbrado", "iluminacion publica deficiente"],
}

# ===================== Utilidades =====================
import re as _re
def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))

def norm_text(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = strip_accents(str(s)).lower().strip()
    s = _re.sub(r"\s+", " ", s)
    return s

def make_unique_columns(cols: List[str]) -> List[str]:
    seen, out = {}, []
    for c in cols:
        nc = norm_text(c)
        seen[nc] = seen.get(nc, 0) + 1
        out.append(nc if seen[nc] == 1 else f"{nc}__{seen[nc]}")
    return out

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = make_unique_columns([str(c) for c in out.columns])
    return out

@st.cache_data(show_spinner=False)
def read_matriz(file_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(BytesIO(file_bytes), sheet_name="matriz", engine="openpyxl")

# --- helpers rango Excel ---
def excel_col_to_index(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

AI_IDX = excel_col_to_index("AI") - 1  # 0-based
ET_IDX = excel_col_to_index("ET") - 1

# --- detectar encabezados amarillos (opcional) ---
def is_yellow(rgb: Optional[str]) -> bool:
    if not rgb: return False
    rgb = str(rgb).upper()
    if len(rgb) == 8: rgb = rgb[2:] if rgb.startswith("FF") else rgb[-6:]
    if len(rgb) != 6: return False
    known = {"FFF2CC","FFEB9C","FFE699","FFFFCC","FFFF00","FFFDE9","FFFCE5"}
    if rgb in known: return True
    try:
        r,g,b = int(rgb[0:2],16), int(rgb[2:4],16), int(rgb[4:6],16)
        return (r>=235 and g>=230 and b<=220)
    except: return False

def yellow_header_indices(file_bytes: bytes, start_col="AI", end_col="ET") -> List[int]:
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb["matriz"]
    c0 = excel_col_to_index(start_col)
    c1 = excel_col_to_index(end_col)
    out = []
    for c in range(c0, c1+1):
        cell = ws.cell(row=1, column=c)
        rgb = None
        if cell.fill:
            rgb = (getattr(cell.fill.fgColor, "rgb", None)
                   or getattr(cell.fill.start_color, "rgb", None))
        if is_yellow(rgb):
            out.append(c-1)  # 0-based para pandas
    return out

# ===================== Conteo celdas con datos =====================
NEGATIVOS = {"", "0", "no", "false", "nan", "ninguno", "n/a", "na"}

def cell_has_data(x) -> bool:
    s = norm_text(x)
    return s not in NEGATIVOS

def header_marked_series(s: pd.Series) -> pd.Series:
    # celdas con "dato": cualquier cosa que no sea vac√≠o/negativo
    return s.apply(cell_has_data)

def counts_from_range(df_all: pd.DataFrame, file_bytes: bytes) -> pd.DataFrame:
    # columnas objetivo AI:ET
    ncols = df_all.shape[1]
    L = max(0, min(AI_IDX, ncols-1))
    R = max(0, min(ET_IDX, ncols-1))
    if L > R:
        L, R = 0, ncols-1

    # intentar filtrar por encabezados amarillos; si no hay, tomar todas AI:ET
    amarillas = [i for i in yellow_header_indices(file_bytes, "AI", "ET") if L <= i <= R]
    idx_cols = amarillas if amarillas else list(range(L, R+1))

    # limitar filas a las que tengan alg√∫n dato en esas columnas (evita filas vac√≠as bajo la tabla)
    df_rng = df_all.iloc[:, idx_cols].copy()
    rowmask = df_rng.applymap(cell_has_data).any(axis=1)
    df_rng = df_rng.loc[rowmask]

    rows = []
    for j, colname in zip(idx_cols, df_all.columns[idx_cols]):
        freq = int(header_marked_series(df_rng[colname]).sum())
        if freq > 0:
            rows.append((str(colname), freq))
    if not rows:
        return pd.DataFrame({"Descriptor": [], "Frecuencia": []})
    return pd.DataFrame(rows, columns=["Descriptor", "Frecuencia"])

# ===================== Canonizaci√≥n / Pareto =====================
def build_canon_maps(dic_df: pd.DataFrame) -> Tuple[Dict[str,str], Dict[str,str]]:
    desc_by_norm, cat_by_norm = {}, {}
    for _, r in dic_df.iterrows():
        d = str(r["Descriptor"]).strip()
        c = _force_cat(str(r["Categor√≠a"]))
        desc_by_norm[norm_text(d)] = d
        cat_by_norm[norm_text(d)]  = c
        for syn in SINONIMOS.get(d, []):
            desc_by_norm[norm_text(syn)] = d
            cat_by_norm[norm_text(syn)]  = c
    return desc_by_norm, cat_by_norm

def canoniza(raw_desc: str, desc_by_norm: Dict[str,str]) -> str:
    n = norm_text(raw_desc)
    if n in desc_by_norm:
        return desc_by_norm[n]
    hit = get_close_matches(n, list(desc_by_norm.keys()), n=1, cutoff=0.82)
    return desc_by_norm[hit[0]] if hit else raw_desc.strip()

def build_pareto(base_counts: pd.DataFrame, dic_df: pd.DataFrame) -> pd.DataFrame:
    if base_counts.empty:
        return pd.DataFrame(columns=["Categor√≠a","Descriptor","Frecuencia","Porcentaje","% acumulado","Acumulado","80/20"])

    desc_by_norm, cat_by_norm = build_canon_maps(dic_df)

    df = base_counts.copy()
    df["Descriptor"] = df["Descriptor"].astype(str)
    df["Descriptor Canon"] = df["Descriptor"].apply(lambda x: canoniza(x, desc_by_norm))
    df["Categor√≠a"] = df["Descriptor Canon"].apply(lambda d: _force_cat(cat_by_norm.get(norm_text(d), "OTROS FACTORES")))

    grp = df.groupby(["Categor√≠a", "Descriptor Canon"], as_index=False)["Frecuencia"].sum()
    grp = grp.rename(columns={"Descriptor Canon": "Descriptor"})

    total = int(grp["Frecuencia"].sum())
    grp = grp.sort_values(["Frecuencia","Descriptor"], ascending=[False, True], ignore_index=True)
    grp["Porcentaje"]  = (grp["Frecuencia"] / total)        # fracci√≥n 0‚Äì1
    grp["% acumulado"] = grp["Porcentaje"].cumsum()          # fracci√≥n 0‚Äì1
    grp["Acumulado"]   = grp["Frecuencia"].cumsum()
    grp["80/20"]       = "80%"

    assert int(grp["Acumulado"].iloc[-1]) == total
    assert abs(float(grp["% acumulado"].iloc[-1]) - 1.0) < 1e-9

    return grp[["Categor√≠a","Descriptor","Frecuencia","Porcentaje","% acumulado","Acumulado","80/20"]]

# ===================== Excel con formato + gr√°fico =====================
def export_excel(pareto: pd.DataFrame, titulo: str = "PARETO COMUNIDAD") -> bytes:
    from pandas import ExcelWriter
    import xlsxwriter  # noqa: F401

    out = BytesIO()
    with ExcelWriter(out, engine="xlsxwriter") as writer:
        sheet = "Pareto Comunidad"
        pareto.to_excel(writer, index=False, sheet_name=sheet)
        wb = writer.book
        ws = writer.sheets[sheet]
        n = len(pareto)
        if not n:
            return out.getvalue()

        fmt_head = wb.add_format({"bold": True, "align": "center", "bg_color": "#D9E1F2", "border": 1})
        fmt_pct  = wb.add_format({"num_format": "0,00%", "align": "right", "border": 1})
        fmt_int  = wb.add_format({"num_format": "#,##0", "align": "center", "border": 1})
        fmt_txt  = wb.add_format({"align": "left", "border": 1})
        fmt_cent = wb.add_format({"align": "center"})
        fmt_yel  = wb.add_format({"bg_color": "#FFF2CC"})

        ws.set_row(0, None, fmt_head)
        ws.set_column("A:A", 22, fmt_txt)
        ws.set_column("B:B", 52, fmt_txt)
        ws.set_column("C:C", 12, fmt_int)
        ws.set_column("D:D", 12, fmt_pct)   # fracci√≥n ‚Üí 0,00%
        ws.set_column("E:E", 12, fmt_pct)
        ws.set_column("F:F", 12, fmt_int)
        ws.set_column("G:G", 8,  fmt_cent)

        cutoff_idx = int((pareto["% acumulado"] <= 0.80).sum())
        if cutoff_idx > 0:
            ws.conditional_format(1, 0, cutoff_idx, 6, {"type": "no_blanks", "format": fmt_yel})

        # auxiliares para las l√≠neas
        ws.write(0, 9, "80/20");  ws.set_column("J:J", 6,  None, {"hidden": True})
        ws.write(0,10, "CorteX"); ws.set_column("K:K", 20, None, {"hidden": True})
        ws.write(0,11, "%");      ws.set_column("L:L", 6,  None, {"hidden": True})
        for i in range(n): ws.write_number(i+1, 9, 0.80)

        corte_row = max(1, cutoff_idx)
        xcat = pareto.iloc[corte_row-1]["Descriptor"]
        ws.write(1,10, xcat); ws.write(2,10, xcat)
        ws.write_number(1,11, 0.0); ws.write_number(2,11, 1.0)

        # Barras
        chart = wb.add_chart({'type': 'column'})
        points = [{"fill": {"color": "#5B9BD5"}} for _ in range(n)]
        for i in range(cutoff_idx, n): points[i] = {"fill": {"color": "#A6A6A6"}}
        chart.add_series({
            'name': 'Frecuencia',
            'categories': [sheet, 1, 1, n, 1],
            'values':     [sheet, 1, 2, n, 2],
            'points': points,
        })

        # % acumulado (eje secundario 0‚Äì100%)
        line = wb.add_chart({'type': 'line'})
        line.add_series({
            'name': '% acumulado',
            'categories': [sheet, 1, 1, n, 1],
            'values':     [sheet, 1, 4, n, 4],
            'y2_axis': True,
            'line': {'color': '#ED7D31', 'width': 2.0}
        })
        chart.combine(line)

        # 80%
        h80 = wb.add_chart({'type': 'line'})
        h80.add_series({
            'name': '80/20',
            'categories': [sheet, 1, 1, n, 1],
            'values':     [sheet, 1, 9, n, 9],
            'y2_axis': True,
            'line': {'color': '#7F7F7F', 'width': 1.25}
        })
        chart.combine(h80)

        # Corte vertical
        vline = wb.add_chart({'type': 'line'})
        vline.add_series({
            'name': '',
            'categories': [sheet, 1, 10, 2, 10],
            'values':     [sheet, 1, 11, 2, 11],
            'y2_axis': True,
            'line': {'color': '#C00000', 'width': 2.25},
            'marker': {'type': 'none'},
        })
        chart.combine(vline)

        chart.set_title({'name': titulo})
        chart.set_plotarea({'border': {'none': True}})
        chart.set_chartarea({'border': {'none': True}})
        chart.set_x_axis({'num_font': {'rotation': -50}})
        chart.set_y_axis({'major_gridlines': {'visible': False}})
        chart.set_y2_axis({'min': 0, 'max': 1, 'major_unit': 0.1, 'num_format': '0%'})
        chart.set_legend({'position': 'bottom'})
        ws.insert_chart(1, 9, chart, {'x_scale': 1.9, 'y_scale': 1.6})

    return out.getvalue()

# ===================== UI =====================
st.title("Pareto Comunidad ‚Äì MSP (DELITO / RIESGO SOCIAL / OTROS FACTORES)")

plantilla = st.file_uploader("üìÑ Sub√≠ la Plantilla (XLSX) ‚Äì hoja `matriz`", type=["xlsx"])
if not plantilla:
    st.info("Sub√≠ la Plantilla para procesar.")
    st.stop()

try:
    df_all = read_matriz(plantilla.getvalue())
except Exception as e:
    st.error(f"Error leyendo 'matriz': {e}")
    st.stop()

df_all = normalize_columns(df_all)
st.caption(f"Vista previa (primeras 20 de {len(df_all)} filas)")
st.dataframe(df_all.head(20), use_container_width=True)

# ---- Conteo en AI:ET de celdas con datos (filtra filas vac√≠as) ----
base = counts_from_range(df_all, plantilla.getvalue())
if base.empty or base["Frecuencia"].sum() == 0:
    st.warning("No se hallaron datos en AI:ET (o no hay columnas v√°lidas). Revis√° la plantilla.")
    st.stop()

# ---- Pareto ----
pareto = build_pareto(base, DICCIONARIO_EMBEBIDO)

# ---- Mostrar con coma decimal en pantalla ----
def pct_str(frac: float) -> str:
    return f"{frac*100:.2f}%".replace(".", ",")

display = pareto.copy()
display["Porcentaje"] = display["Porcentaje"].apply(pct_str)
display["% acumulado"] = display["% acumulado"].apply(pct_str)

TOTAL = int(pareto["Acumulado"].iloc[-1])
st.subheader(f"Pareto Comunidad (TOTAL = {TOTAL:,})")
st.dataframe(display, use_container_width=True)

# ---- Gr√°fico r√°pido (0‚Äì100% en pantalla) ----
import altair as alt
top_df = pareto.head(TOP_N_GRAFICO).copy()
bars = alt.Chart(top_df).mark_bar().encode(
    x=alt.X('Descriptor:N', sort=None, axis=alt.Axis(labelAngle=-50)),
    y=alt.Y('Frecuencia:Q')
)
line = alt.Chart(top_df).mark_line(point=True).encode(
    x='Descriptor:N',
    y=alt.Y('% acumulado:Q', axis=alt.Axis(format='%'), scale=alt.Scale(domain=[0,1])),
    color=alt.value('#ED7D31')
)
h80 = alt.Chart(pd.DataFrame({'y':[0.8]})).mark_rule().encode(y=alt.Y('y:Q', axis=alt.Axis(format='%')))
st.altair_chart((bars + line + h80).resolve_scale(y='independent'), use_container_width=True)

# ---- Descargar Excel ----
st.subheader("Descargar Excel final")
st.download_button(
    "‚¨áÔ∏è Pareto Comunidad (Excel con formato y gr√°fico)",
    data=export_excel(pareto, titulo="PARETO COMUNIDAD"),
    file_name="Pareto_Comunidad.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)




